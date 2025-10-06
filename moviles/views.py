from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import render_to_string
from rest_framework.response import Response
from django.db.models import Sum
import django.db.models as models
from django.http import JsonResponse
from datetime import datetime, timedelta
from rest_framework.generics import ListAPIView

# Create your views here.
from moviles.serializer import CargaCombustibleSerializer, CentroCostosSerializer, GastosMovilSerializer, MatafuegosSerializer, MovilSerializer, PersonalSerializer, TipoVencimientosSerializer, VencimientosSerializer
from rest_framework import filters

from syh.models import Area, Cliente
from utiles.BaseViewSet import BaseAppModelViewSet
from .models import CargaCombustible, CentroCostos, GastosMovil, Matafuegos, Movil, Personal, TipoVencimientos, Vencimientos
from .models import Viajes, Predios
from .serializer import ViajesSerializer, PrediosSerializer
from rest_framework.decorators import api_view
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter


class MovilViewSet(BaseAppModelViewSet):
    queryset = Movil.objects.all()
    serializer_class = MovilSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['patente', 'empresa__nombre']  # Asegúrate de que 'nombre' es el campo correcto en el modelo 'Clientes'
    ordering_fields = ['patente']
    

class TipoVencimientosViewSet(BaseAppModelViewSet):
    queryset = TipoVencimientos.objects.all()
    serializer_class = TipoVencimientosSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre', 'tipo']  # Asegúrate de que 'nombre' y 'tipo' son los campos correctos en el modelo 'TipoVencimientos'
    ordering_fields = ['nombre']

class VencimientosViewSet(BaseAppModelViewSet):
    queryset = Vencimientos.objects.all()
    serializer_class = VencimientosSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['movil__patente', 'tipo_vencimiento__nombre', 'detalle']  # Asegúrate de que 'patente', 'nombre' y 'detalle' son los campos correctos en los modelos 'Movil' y 'TipoVencimientos'
    ordering_fields = ['-fecha', 'tipo_vencimiento']

class PersonalViewSet(BaseAppModelViewSet):
    queryset = Personal.objects.all()
    serializer_class = PersonalSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre', 'apellido', 'empresa__nombre']  # Asegúrate de que 'nombre', 'apellido' y 'dni' son los campos correctos en el modelo 'Personal'
    ordering_fields = ['nombre', 'apellido']
    
class CargaCombustibleViewSet(BaseAppModelViewSet):
    queryset = CargaCombustible.objects.all()
    serializer_class = CargaCombustibleSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['movil__patente', 'chofer__nombre', 'chofer__apellido', 'empresa__nombre']
    ordering_fields = ['fecha']

class MatafuegosViewSet(BaseAppModelViewSet):
    queryset = Matafuegos.objects.all()
    serializer_class = MatafuegosSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['empresa__nombre', 'area_nombre']
    ordering_fields = ['fecha_vencimiento']

class GastosMovilViewSet(BaseAppModelViewSet):
    queryset = GastosMovil.objects.all()
    serializer_class = GastosMovilSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['movil__patente', 'cliente__nombre', 'area__detalle', 'centro_costos__descripcion']
    ordering_fields = ['fecha']

class CentroCostosViewSet(BaseAppModelViewSet):
    queryset = CentroCostos.objects.all()
    serializer_class = CentroCostosSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['descripcion']
    ordering_fields = ['descripcion']


class PrediosViewSet(BaseAppModelViewSet):
    queryset = Predios.objects.all()
    serializer_class = PrediosSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre']
    ordering_fields = ['nombre']


class ViajesViewSet(BaseAppModelViewSet):
    queryset = Viajes.objects.all()
    serializer_class = ViajesSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['movil__patente', 'cliente__nombre', 'origen__nombre', 'destino', 'producto']
    ordering_fields = ['-fecha']
    ordering = ['-fecha', '-id']  # Ordenar por fecha descendente y luego por ID

    def get_queryset(self):
        qs = super().get_queryset()
        # filtros opcionales por rango de fecha y movil
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        movil = self.request.query_params.get('movil')
        chofer = self.request.query_params.get('chofer')
        producto = self.request.query_params.get('producto')
        destino = self.request.query_params.get('destino')
        if start_date:
            qs = qs.filter(fecha__gte=start_date)
        if end_date:
            qs = qs.filter(fecha__lte=end_date)
        if movil:
            qs = qs.filter(movil_id=movil)
        if chofer:
            # filtrar por personal/chofer (campo personal en Viajes)
            qs = qs.filter(personal_id=chofer)
        if producto:
            qs = qs.filter(producto=producto)
        if destino:
            qs = qs.filter(destino=destino)
        return qs.order_by('-fecha', '-id')  # Asegurar orden consistente

    # endpoint extra para KPIs
    def kpis(self, request):
        qs = self.filter_queryset(self.get_queryset())
        total_viajes = qs.count()
        tn_pulpable = qs.aggregate(sum=Sum('tn_pulpable'))['sum'] or 0
        tn_aserrable = qs.aggregate(sum=Sum('tn_aserrable'))['sum'] or 0
        tn_chip = qs.aggregate(sum=Sum('tn_chip'))['sum'] or 0
        totals = {
            'total_viajes': total_viajes,
            'tn_pulpable': float(tn_pulpable),
            'tn_aserrable': float(tn_aserrable),
            'tn_chip': float(tn_chip),
            'tn_total': float(tn_pulpable + tn_aserrable + tn_chip)
        }
        return Response(totals)
    
def moviles_list(request):
    return render(request, 'pages/moviles/movil_list.html')

def tipo_vencimientos_list(request):
    return render(request, 'pages/moviles/tipo_vencimientos_list.html')

def vencimientos_list(request):
    return render(request, 'pages/moviles/vencimientos_list.html')

def personal_list(request):
    return render(request, 'pages/moviles/personal_list.html')

def carga_combustible_list(request):
    return render(request, 'pages/moviles/carga_combustible_list.html')

def gastos_movil_list(request):
    return render(request, 'pages/moviles/gastos_moviles_list.html')

def consumo_combustible_por_mes(request):
    empresas = Cliente.objects.all()
    fecha_inicio = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
    fecha_fin = datetime.now().strftime('%Y-%m-%d')
    return render(request, 
                  'pages/moviles/consumo_combustible_por_mes.html', 
                  {'empresas': empresas, 'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin})

def datos_consumo_combustible(request):
    empresa_id = request.GET.get('empresa')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    if not fecha_inicio:
        fecha_inicio = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
    if not fecha_fin:
        fecha_fin = datetime.now().strftime('%Y-%m-%d')
    if empresa_id:
        empresa = Cliente.objects.get(id=empresa_id)
        datos = CargaCombustible.objects.filter(empresa=empresa, fecha__range=[fecha_inicio, fecha_fin]).values('movil__patente', 'fecha__month').annotate(consumo=Sum('litros'))
        datos_json = []
        for dato in datos:
            datos_json.append({
              'movil': dato['movil__patente'],
              'mes': dato['fecha__month'],
                'consumo': dato['consumo']
            })
        return JsonResponse(datos_json, safe=False)
    else:
        return JsonResponse([], safe=False)
    

#vistas para el manejo de matafuegos
def matafuegos_list(request):
    return render(request, 'pages/moviles/matafuegos_list.html')

class GastosPorMovilView(ListAPIView):
    serializer_class = GastosMovilSerializer
    
    def get_queryset(self):
        queryset = GastosMovil.objects.all()
        
        # Filtros
        empresa = self.request.query_params.get('empresa')
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        
        if empresa:
            queryset = queryset.filter(cliente_id__id=empresa)
        if start_date:
            queryset = queryset.filter(fecha__gte=start_date)
        if end_date:
            queryset = queryset.filter(fecha__lte=end_date)
            
        return queryset
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        
        # Datos para la tabla
        serializer = self.get_serializer(queryset, many=True)
        
        # Datos para el gráfico de torta
        chart_data = queryset.values('centro_costos__descripcion') \
            .annotate(total_importe=Sum('importe')) \
            .order_by('-total_importe')
        
        return Response({
            'gastos': serializer.data,
            'chart_data': list(chart_data)
        })


def reporte_gastos_view(request):
    empresas = Cliente.objects.all()
    # Establecer fechas por defecto (último mes)
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=30)
    return render(request, 'pages/moviles/reporte_gastos_movil.html', {
        'empresas': empresas,
        'start_date': start_date,
        'end_date': end_date
    })        


def viajes_list(request):
    # página que carga la app Vue para manejar viajes
    return render(request, 'pages/moviles/viajes_list.html')


@api_view(['GET'])
def viajes_kpis(request):
    """Endpoint simple que devuelve KPIs de viajes filtrando por start_date, end_date, movil"""
    qs = Viajes.objects.all()
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    movil = request.query_params.get('movil')
    chofer = request.query_params.get('chofer')
    producto = request.query_params.get('producto')
    destino = request.query_params.get('destino')
    if start_date:
        qs = qs.filter(fecha__gte=start_date)
    if end_date:
        qs = qs.filter(fecha__lte=end_date)
    if movil:
        qs = qs.filter(movil_id=movil)
    if chofer:
        qs = qs.filter(personal_id=chofer)
    if producto:
        qs = qs.filter(producto=producto)
    if destino:
        qs = qs.filter(destino=destino)

    total_viajes = qs.count()
    tn_pulpable = qs.aggregate(sum=Sum('tn_pulpable'))['sum'] or 0
    tn_aserrable = qs.aggregate(sum=Sum('tn_aserrable'))['sum'] or 0
    tn_chip = qs.aggregate(sum=Sum('tn_chip'))['sum'] or 0
    totals = {
        'total_viajes': total_viajes,
        'tn_pulpable': float(tn_pulpable),
        'tn_aserrable': float(tn_aserrable),
        'tn_chip': float(tn_chip),
        'tn_total': float(tn_pulpable + tn_aserrable + tn_chip)
    }
    return Response(totals)


@api_view(['GET'])
def export_viajes_xlsx(request):
    """Exporta los viajes filtrados a un archivo XLSX.

    Aplica los mismos filtros que el endpoint de KPIs/get_queryset: start_date, end_date, movil, chofer, producto, destino
    Incluye viajes con sin_actividad=True incluso si origen es NULL.
    """
    # Usar select_related solo en movil (obligatorio) y personal (nullable)
    # No usar select_related en origen para evitar problemas con NULL en viajes sin actividad
    qs = Viajes.objects.all()
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    movil = request.query_params.get('movil')
    chofer = request.query_params.get('chofer')
    producto = request.query_params.get('producto')
    destino = request.query_params.get('destino')
    if start_date:
        qs = qs.filter(fecha__gte=start_date)
    if end_date:
        qs = qs.filter(fecha__lte=end_date)
    if movil:
        qs = qs.filter(movil_id=movil)
    if chofer:
        qs = qs.filter(personal_id=chofer)
    if producto:
        qs = qs.filter(producto=producto)
    if destino:
        qs = qs.filter(destino=destino)

    # Ordenar y contar para debug
    qs = qs.order_by('-fecha', '-id')
    total_viajes = qs.count()
    viajes_sin_actividad = qs.filter(sin_actividad=True).count()
    print(f"[EXPORT] Total viajes a exportar: {total_viajes}, con sin_actividad: {viajes_sin_actividad}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Viajes'

    # Construir encabezados tomando verbose_name cuando esté disponible
    model = Viajes

    def label_for(field_name):
        try:
            f = model._meta.get_field(field_name)
            return str(f.verbose_name).title()
        except Exception:
            return field_name.replace('_', ' ').title()


    # Sólo las columnas requeridas por el usuario, en el orden solicitado
    headers = [
        ('fecha', 'Fecha'),
        ('movil_patente', 'Patente'),
        ('personal_nombre', 'Nombre Chofer'),
        ('origen', 'Origen'),
        ('origen_nombre', 'Origen Nombre'),
        ('destino', 'Destino'),
        ('producto', 'Producto'),
        ('tn_pulpable', 'KG Pulpable'),
        ('tn_aserrable', 'KG Aserrable'),
        ('tn_chip', 'KG Chips'),
        ('sin_actividad', 'Sin Actividad'),
        ('motivo_sin_actividad', 'Motivo Sin Actividad'),
        ('observaciones', 'Observaciones'),
    ]

    ws.append([h[1] for h in headers])

    # Iterar y exportar, manejando casos donde origen puede ser NULL
    rows_exported = 0
    for obj in qs:
        fecha = obj.fecha.strftime('%d-%m-%Y') if getattr(obj, 'fecha', None) else ''
        movil_patente = obj.movil.patente if obj.movil else ''
        personal_nombre = f"{obj.personal.apellido} {obj.personal.nombre}" if obj.personal else ''
        # Manejar origen NULL (puede ocurrir en viajes sin actividad)
        try:
            origen_val = obj.origen.id if obj.origen else ''
            origen_nombre = obj.origen.nombre if obj.origen else ''
        except Exception:
            origen_val = ''
            origen_nombre = ''
        destino = obj.get_destino_display() if hasattr(obj, 'get_destino_display') else (obj.destino or '')
        producto = obj.get_producto_display() if hasattr(obj, 'get_producto_display') else (obj.producto or '')
        tn_pulpable = float(obj.tn_pulpable) if obj.tn_pulpable is not None else 0
        tn_aserrable = float(obj.tn_aserrable) if obj.tn_aserrable is not None else 0
        tn_chip = float(obj.tn_chip) if obj.tn_chip is not None else 0
        sin_actividad = bool(obj.sin_actividad)
        motivo = obj.motivo_sin_actividad or ''
        observaciones = obj.observaciones or ''

        row = [
            fecha,
            movil_patente,
            personal_nombre,
            origen_val,
            origen_nombre,
            destino,
            producto,
            tn_pulpable,
            tn_aserrable,
            tn_chip,
            sin_actividad,
            motivo,
            observaciones,
        ]
        ws.append(row)
        rows_exported += 1

    print(f"[EXPORT] Filas exportadas: {rows_exported}")

    # Ajustar anchos de columna
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            try:
                v = str(cell.value) if cell.value is not None else ''
            except Exception:
                v = ''
            if len(v) > max_length:
                max_length = len(v)
        ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"viajes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response